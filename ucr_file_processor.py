import pandas as pd
import numpy as np
from sqlalchemy import create_engine
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
import os.path
import numpy as np
import sys



base_path = Path(__file__).parent
#All files used
RAW_DATA_FILE_PATH= (base_path/ "ucr_raw_data_files/").resolve()
RAW_ICD_DATA_FILE_PATH= (base_path/ "ICD_raw_data_files/").resolve()
OUTPUT_DATA_FILE_PATH= (base_path/ "ucr_output_data_files/").resolve()
MIT_PROCESSED_FILES=(base_path/ "mit_mortality_published_files/").resolve()
RAW_CDC_POPULATION_FILES=(base_path/ "population_data_files/").resolve()
POPULATION_FILES=(base_path/ "population_data_files/").resolve()
DECADES_MORTALITY_DATA_TXT_FILES=(base_path/ "mortality_txt_files/").resolve()


MAPPING_DATA_FILE_PATH=(base_path).resolve()

engine = create_engine('sqlite:///mit_mortality_data.db', echo=False)
sqlite_connection = engine.connect()

#SQLite tables used
RAW_MORTALITY_DATA_TABLE = "CANCER_MORTALITY_RAW_DATA"
ICD10_RAW_MORTALITY_DATA_TABLE = "ICD10_CANCER_MORTALITY_RAW_DATA"
UCR358_CODES_TABLE="UCR358_CODE_DESCRIPTION"
MIT_UCR358_CODE_MAPPING_TABLE="MIT_UCR358_ICD10_CODE_TRANSLATION"
MORTALITY_BY_BIRTH_YEAR_TABLE="MIT_MORTALITY_BY_BIRTH_YEAR_DATA"

#Population tables
POPULATION_DATA_TABLE="AGE_SEX_RACE_POPULATION"
RAW_CDC_POPULATION_DATA_TABLE="CDC_POPULATION"
POPULATION_TOT_ADJ_DATA_TABLE="TOT_ADJUSTMENT_POPULATION"
RAW_CDC_85_PLUS_POPULATION_DATA_TABLE="CDC_85_PLUS_POPULATION"

#this is used to store the existing data from the excel files as a backup and testing
ORIG_MORTALITY_DATA_TABLE="ORIGINAL_MIT_RAW_CANCER_MORTALITY_DATA"
ORIG_1_MINUS_TOT_DATA_TABLE="ORIGINAL_MIT_1_MINUS_TOT_DATA"
ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE="ORIGINAL_MIT_RAW_ADJUSTED_MORTALITY_DATA"
ORIG_MORTALITY_BY_BIRTH_YEAR_TABLE="ORIGINAL_MIT_MORTALITY_BY_BIRTH_YEAR_DATA"


# This table gas data after mapping UCR358 codes to MIT category
MIT_RAW_MORTALITY_DATA_TABLE = "CANCER_MORTALITY_MIT_RAW_DATA"



# AGE_OF_DEATH_ORDER= ['under one year',	'1 year',	'2 years',	'3 years',	'4 years','under 5 years','5-9 years',	'10-14 years',	'15-19 years',	'20-24 years',	'25-29 years',	'30-34 years',	'35-39 years ',	'40-44 years',	'45-49 years',	'50-54 years ',	'55-59 years',	'60-64 years',	'65-69 years',	'70-74 years',	'75-79 years',	'80-84 years',	'85-89 years',	'90-94 years',	'95-99 years',	'100+ years',	'Not stated']

AGE_OF_DEATH_ORDER= ['under_one_year',	'year_1',	'years_2',	'years_3',	'years_4','under_5_years','years_5_9',	'years_10_14',	'years_15_19',	'years_20_24',	'years_25_29',	'years_30_34',	'years_35_39',	'years_40_44',	'years_45_49',	'years_50_54',	'years_55_59',	'years_60_64',	'years_65_69',	'years_70_74',	'years_75_79',	'years_80_84',	'years_85_89',	'years_90_94',	'years_95_99',	'years_100_plus',	'Not_stated']

#remove this so that we do not duplicate total counts
# AGE_OF_DEATH_SUM_COLS=[age for age in AGE_OF_DEATH_ORDER if age != 'under 5 years']
AGE_OF_DEATH_SUM_COLS=[age for age in AGE_OF_DEATH_ORDER if age != 'under_5_years']

AGE_OF_DEATH_POPULATION_COLS=[age for age in AGE_OF_DEATH_ORDER if age not in ['under_one_year',	'year_1',	'years_2',	'years_3',	'years_4','under_5_years','Not_stated']]
COL_ORDER_DECADES = ['AGE_AT_DEATH', 'ETHNICITY', 1800, 1810, 1820, 1830, 1840, 1850,
                     1860, 1870, 1880, 1890, 1900, 1910, 1920, 1930,
                     1940, 1950, 1960, 1970, 1980, 1990, 2000, 2010
                     ]

#mapping dictionaries
AGE_27_RECODE_MAPPING={
2:'under one year',
3:'1 year',
4:'2 years',
5:'3 years',
6:'4 years',
7:'5-9 years',
8:'10-14 years',
9:'15-19 years',
10:'20-24 years',
11:'25-29 years',
12:'30-34 years',
13:'35-39 years',
14:'40-44 years',
15:'45-49 years',
16:'50-54 years',
17:'55-59 years',
18:'60-64 years',
19:'65-69 years',
20:'70-74 years',
21:'75-79 years',
22:'80-84 years',
23:'85-89 years',
24:'90-94 years',
25:'95-99 years',
26:'100+ years',
27:'Not stated'

}
ethnicity_dict_2011={
    'White Male':'EAM',
    'White Female':'EAF',
    'Non-White Male':'NEAM',
    'Non-White Female':'NEAF',

    }

ethnicity_dict={
    'WHITE_MALES':'EAM',
    'WHITE_FEMALES':'EAF',
    'NON-WHITE_MALES':'NEAM',
    'NON-WHITE_FEMALES':'NEAF',
    'WHITE_MALE': 'EAM',
    'WHITE_FEMALE': 'EAF',
    'NON-WHITE_MALE': 'NEAM',
    'NON-WHITE_FEMALE': 'NEAF'
    }

age_of_death_dict_for_charts={
'under_one_year':0.5,
# 'years_1_4':3,
# 'year_1':3,
# 'years_2':3,
# 'years_3':3,
# 'years_4':3,
'year_1':1,
'years_2':2,
'years_3':3,
'years_4':4,
'years_5_9':7.5,
'years_10_14':12.5,
'years_15_19':17.5,
'years_20_24':22.5,
'years_25_29':27.5,
'years_30_34':32.5,
'years_35_39':37.5,
'years_40_44':42.5,
'years_45_49':47.5,
'years_50_54':52.5,
'years_55_59':57.5,
'years_60_64':62.5,
'years_65_69':67.5,
'years_70_74':72.5,
'years_75_79':77.5,
'years_80_84':82.5,
'years_85_89':87.5,
'years_90_94':92.5,
'years_95_99':97.5,
'years_100_plus':102.5
}

age_of_death_dict={
'under_one_year':0.5,
'years_1_4':3,
'year_1':3,
'years_2':3,
'years_3':3,
'years_4':3,
# 'year_1':1,
# 'years_2':2,
# 'years_3':3,
# 'years_4':4,
'years_5_9':7.5,
'years_10_14':12.5,
'years_15_19':17.5,
'years_20_24':22.5,
'years_25_29':27.5,
'years_30_34':32.5,
'years_35_39':37.5,
'years_40_44':42.5,
'years_45_49':47.5,
'years_50_54':52.5,
'years_55_59':57.5,
'years_60_64':62.5,
'years_65_69':67.5,
'years_70_74':72.5,
'years_75_79':77.5,
'years_80_84':82.5,
'years_85_89':87.5,
'years_90_94':92.5,
'years_95_99':97.5,
'years_100_plus':102.5
}

cdc_85_plus_race_sex_dict={
    1:'EAM',
    2:'EAF',
    3:'NEAM',
    4:'NEAF',
    5:'NEAM',
    6:'NEAF',
    7:'NEAM',
    8:'NEAF'
}

MALE_ONLY_CANCER = ['Male Breast Cancer', 'Testicular and Penile Cancer', 'Testicular Cancer', 'Prostate Cancer', 'Penis and other Male Genital Organs', 'Penile Cancer']
FEMALE_ONLY_CANCER = ['Female Genital Cancer-Other', 'Ovarian Cancer', 'Female Breast Cancer', 'Uteri', 'Cervix Uteri', 'Uteri-Other', 'BN of Uterus']


#=========Load raw UCR358 data in to  table
def load_ucr_files(start_year=2025):
    directory = os.fsencode(RAW_DATA_FILE_PATH)

    for file in os.listdir(directory):
        filename = os.fsdecode(file)
        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            file = os.path.join(RAW_DATA_FILE_PATH, filename)
            ucr_parser = pd.ExcelFile(file)
            if(filename[0].isdigit()):
                year = filename.split(' ')[0]
                ethnicity_in_file=''
                eth_dict=ethnicity_dict_2011
            else: #filename is of format "White_Females_2015.xlx"
                file_str_splits=filename.split('_') #['White','Females','2015.xlx']
                year = file_str_splits.pop().split('.')[0] #2015
                ethnicity_in_file=(file_str_splits[0].upper() + '_' + file_str_splits[1].upper()) #'WHITE_FEAMLES
                eth_dict = ethnicity_dict

            year=int(year)

            if(year<start_year): # we are not going to load the data prior to the start_year
                continue

            for ethnicity in eth_dict:
                if(len(ethnicity_in_file)>1):
                    if(ethnicity==ethnicity_in_file):

                        df_eam = ucr_parser.parse('Sheet1', skiprows=38, usecols="B:AB",header=None)
                    else:
                        continue;
                else:
                    df_eam = ucr_parser.parse(ethnicity, skiprows=39, usecols="B:AB",header=None)

                COLUMN_ORDER = ['UCR358_CODE']
                COLUMN_ORDER.extend(AGE_OF_DEATH_SUM_COLS)
                # df_eam.columns = ['UCR358_CODE', 'under one year',	'1 year',	'2 years',	'3 years',	'4 years',	'5-9 years',	'10-14 years',	'15-19 years',	'20-24 years',	'25-29 years',	'30-34 years',	'35-39 years ',	'40-44 years',	'45-49 years',	'50-54 years ',	'55-59 years',	'60-64 years',	'65-69 years',	'70-74 years',	'75-79 years',	'80-84 years',	'85-89 years',	'90-94 years',	'95-99 years',	'100+ years',	'Not stated']
                df_eam.columns=COLUMN_ORDER
                df_eam['YEAR']=year
                df_eam['ETHNICITY']=eth_dict.get(ethnicity)
                df_eam=df_eam.dropna(how='all')
                # this is extra age category

                df_eam['under_5_years'] = df_eam[['under_one_year', 'year_1', 'years_2', 'years_3', 'years_4']].sum(axis=1)

                df_eam['TOTAL'] = df_eam[AGE_OF_DEATH_SUM_COLS].sum(axis=1)
                df_eam = df_eam.dropna(subset=['UCR358_CODE'])

                #delete if it already exists for that year
                insert_raw_mortality_data(RAW_MORTALITY_DATA_TABLE,df_eam,eth_dict.get(ethnicity),start_year)
                # df_eam.to_sql(RAW_MORTALITY_DATA_TABLE, sqlite_connection, if_exists='append',index=False)
                print(" Loaded Data for: Year",year," ETHNICITY:",eth_dict.get(ethnicity))

def insert_raw_mortality_data(table_name,df_data,ethnicity,start_year):

    #check if table exists
    df_table = pd.read_sql_query(
        "SELECT  name  from sqlite_master where type='table' AND name='"+table_name+"'",
        sqlite_connection)
    if(df_table.shape[0]>0):
        df = pd.read_sql_query(
            "SELECT *  from " + table_name +" WHERE YEAR="+str(start_year)+" and ETHNICITY='"+ethnicity+"'",
            sqlite_connection)
        if(df.shape[0]>0): # if the dat is already there then remove it
            print('Deleting existing Data for Year ',start_year,' from ',table_name)

            sql = "DELETE FROM "+table_name+" WHERE YEAR>="+str(start_year) +" and ETHNICITY='"+ethnicity+'"'

            sqlite_connection.execute(sql)

    #append the new data
    df_data.to_sql(table_name, sqlite_connection, if_exists='append', index=False)

#=================================================

def load_ICD10_files(start_year=2011):
    directory = os.fsencode(RAW_ICD_DATA_FILE_PATH)

    for file in os.listdir(directory):
        filename = os.fsdecode(file)
        print("Processing File :",filename)
        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            file = os.path.join(RAW_ICD_DATA_FILE_PATH, filename)
            ucr_parser = pd.ExcelFile(file)
            if(filename[0].isdigit()):
                year = filename.split(' ')[0]
                ethnicity_in_file=''
                eth_dict=ethnicity_dict_2011
            else: #filename is of format "White_Females_2015.xlx"
                file_str_splits=filename.split('_') #['White','Females','2015.xlx']
                year = file_str_splits.pop().split('.')[0] #2015
                ethnicity_in_file=(file_str_splits[0].upper() + '_' + file_str_splits[1].upper()) #'WHITE_FEAMLES
                eth_dict = ethnicity_dict

            year=int(year)

            if(year<start_year): # we are not going to load the data prior to the start_year
                continue

            for ethnicity in eth_dict:
                if(len(ethnicity_in_file)>1):
                    if(ethnicity==ethnicity_in_file):

                        df_eam = ucr_parser.parse('Sheet1', skiprows=38, usecols="B:AB",header=None)
                    else:
                        continue;
                else:
                    df_eam = ucr_parser.parse(ethnicity, skiprows=39, usecols="B:AB",header=None)

                COLUMN_ORDER = ['ICD10_CODE']
                COLUMN_ORDER.extend(AGE_OF_DEATH_SUM_COLS)
                # df_eam.columns = ['UCR358_CODE', 'under one year',	'1 year',	'2 years',	'3 years',	'4 years',	'5-9 years',	'10-14 years',	'15-19 years',	'20-24 years',	'25-29 years',	'30-34 years',	'35-39 years ',	'40-44 years',	'45-49 years',	'50-54 years ',	'55-59 years',	'60-64 years',	'65-69 years',	'70-74 years',	'75-79 years',	'80-84 years',	'85-89 years',	'90-94 years',	'95-99 years',	'100+ years',	'Not stated']
                df_eam.columns=COLUMN_ORDER
                df_eam['YEAR']=year
                df_eam['ETHNICITY']=eth_dict.get(ethnicity)
                df_eam=df_eam.dropna(how='all')
                # this is extra age category
                # df_eam['under_5_years']=df_eam['under_one_year']+df_eam['year_1']+df_eam['years_2']+df_eam['years_3']+df_eam['years_4']
                df_eam['under_5_years'] = df_eam[['under_one_year', 'year_1', 'years_2', 'years_3', 'years_4']].sum(axis=1)


                df_eam['TOTAL'] = df_eam[AGE_OF_DEATH_SUM_COLS].sum(axis=1)
                df_eam = df_eam.dropna(subset=['ICD10_CODE'])

                #delete if it already exists for that year
                insert_raw_mortality_data(ICD10_RAW_MORTALITY_DATA_TABLE,df_eam,eth_dict.get(ethnicity),year)
                # df_eam.to_sql(ICD10_RAW_MORTALITY_DATA_TABLE, sqlite_connection, if_exists='append',index=False)
                print(" Loaded ICD10 Data for: Year",year," ETHNICITY:",eth_dict.get(ethnicity))


def load_ucr358_codes(sheetname,tablename):
    filename="UCR_358_mapping_updated_v2.xlsx"
    file = os.path.join(MAPPING_DATA_FILE_PATH, filename)
    xl_parser = pd.ExcelFile(file)
    df=xl_parser.parse(sheetname)
    df=df.dropna(how='all')

    if(sheetname=='2011_UCR358_CODES'): # This mapping could be specific for a year
        df['YEAR']=sheetname.split('_')[0]

    df.to_sql(tablename, sqlite_connection, if_exists='replace')
    print("Loaded UCR358 code translation in table:",tablename)


def get_ucr358_codes(year=2011):
    df = pd.read_sql_query(
        "SELECT distinct UCR358_CODE,DESCRIPTION from " + UCR358_CODES_TABLE + " where YEAR=" + str(year),
        sqlite_connection)
    return df;



def get_mit_ucr358_mapping():
    df = pd.read_sql_query(
        "SELECT distinct MIT_DISEASES_CATEGORY,UCR358_CODE,ICD10_CODE,MIT_DATA_FILE_NAME,UPDATE_RAW_DATA_YEAR,UPDATE_POPULATION_YEAR,UPDATE_TOT_ADJ_YEAR,UPDATE_1_MINUS_TOT_ADJ_YEAR  from " + MIT_UCR358_CODE_MAPPING_TABLE,
        sqlite_connection)
    return df;

def get_ucr358_raw_data(ucr358_list=None,code_type='UCR358_CODE'):

    if(code_type=='UCR358_CODE'):
        table_name=RAW_MORTALITY_DATA_TABLE
    else:
        table_name=ICD10_RAW_MORTALITY_DATA_TABLE



    # if(ucr358_list==None):
    #     query="SELECT * from " + RAW_MORTALITY_DATA_TABLE+" group by YEAR,UCR358_CODE,ETHNICITY,TOTAL"
    # else:
        # query = "SELECT * from " + RAW_MORTALITY_DATA_TABLE+" where UCR358_CODE in ("+str(ucr358_list)+")group by YEAR,UCR358_CODE,ETHNICITY,TOTAL"
    query ="""
        select
        YEAR, ETHNICITY,
        SUM(under_one_year) as under_one_year,
        SUM(year_1) as year_1,
        SUM(years_2) as years_2,
        SUM(years_3) as years_3,
        SUM(years_4) as years_4,
        SUM(under_5_years) as under_5_years,
        SUM(years_5_9) as years_5_9,
        SUM(years_10_14) as years_10_14,
        SUM(years_15_19) as years_15_19,
        SUM(years_20_24) as years_20_24,
        SUM(years_25_29) as years_25_29,
        SUM(years_30_34) as years_30_34,
        SUM(years_35_39) as years_35_39,
        SUM(years_40_44) as years_40_44,
        SUM(years_45_49) as years_45_49,
        SUM(years_50_54) as years_50_54,
        SUM(years_55_59) as years_55_59,
        SUM(years_60_64) as years_60_64,
        SUM(years_65_69) as years_65_69,
        SUM(years_70_74) as years_70_74,
        SUM(years_75_79) as years_75_79,
        SUM(years_80_84) as years_80_84,
        SUM(years_85_89) as years_85_89,
        SUM(years_90_94) as years_90_94,
        SUM(years_95_99) as years_95_99,
        SUM(years_100_plus) as years_100_plus,
        SUM(Not_stated) as Not_stated,
        SUM(TOTAL) as TOTAL
        from #TABLE_NAME#
        where
        #CODE_TYPE# in (#UCR_CODE_LIST#)
        group
        by
        YEAR, ETHNICITY
        """

    query=query.replace('#UCR_CODE_LIST#',str(ucr358_list))
    query = query.replace('#TABLE_NAME#', table_name)
    query = query.replace('#CODE_TYPE#', code_type)

    df = pd.read_sql_query(
        query,
        sqlite_connection)
    return df




#==============This groups the MIT ucr358codes as per the config file and updates all excel files sheets/inserts in to sqlite table

# ======================== Load the existing data from original files=================
def load_existing_mortality_raw_data(file_path,cancer_type_in=None):

    DATA_FILE_PATH=file_path
    print("Loading Data from file:",DATA_FILE_PATH)
    # read all files and append the data frame
    directory = os.fsencode(DATA_FILE_PATH)

    for file in os.listdir(directory):
        filename = os.fsdecode(file)
        if filename.endswith(".xls") or filename.endswith(".xlsx"):
            file = os.path.join(DATA_FILE_PATH, filename)
            # cancer_xl = pd.ExcelFile(file)
            cancer_xl=None
            cancer_type = filename.split('.')[0]
            cancer_type = cancer_type.replace('Cancer of the', '').replace('Cancer of', '').replace('_06', '')
            cancer_type = cancer_type.strip()

            if(cancer_type in ['mortality_data_comparison']):
                print("Skipping ",cancer_type)
                continue;
            if(cancer_type_in is not None):
                if(cancer_type!=cancer_type_in):
                    continue



            # if(cancer_type not in ['Female Genital Cancer-Other','Ovarian Cancer','Female Breast Cancer']): # EAM/NEAM sheets are not present in the files for these two cancer types

            COLUMN_ORDER = ['YEAR', 'TOTAL']
            COLUMN_ORDER.extend(AGE_OF_DEATH_ORDER)
            try:
                save_original_data(file,'(EAM) MOR(t)','EAM',cancer_type,COLUMN_ORDER,"A:AC",ORIG_MORTALITY_DATA_TABLE)
                save_original_data(file, '(NEAM) MOR(t)', 'NEAM', cancer_type,COLUMN_ORDER,"A:AC",ORIG_MORTALITY_DATA_TABLE)
                save_original_data(file, '(EAF) MOR(t)', 'EAF', cancer_type, COLUMN_ORDER, "A:AC", ORIG_MORTALITY_DATA_TABLE)
                save_original_data(file, '(NEAF) MOR(t)', 'NEAF', cancer_type, COLUMN_ORDER, "A:AC", ORIG_MORTALITY_DATA_TABLE)
            except Exception:
                save_original_data(file,'Raw Data (EAM)','EAM',cancer_type,COLUMN_ORDER,"A:AC",ORIG_MORTALITY_DATA_TABLE)
                save_original_data(file, 'Raw Data (NEAM)', 'NEAM', cancer_type,COLUMN_ORDER,"A:AC",ORIG_MORTALITY_DATA_TABLE)
                save_original_data(file, 'Raw Data (EAF)', 'EAF', cancer_type, COLUMN_ORDER, "A:AC", ORIG_MORTALITY_DATA_TABLE)
                save_original_data(file, 'Raw Data (NEAF)', 'NEAF', cancer_type, COLUMN_ORDER, "A:AC", ORIG_MORTALITY_DATA_TABLE)



            # COLUMN_ORDER = ['YEAR', 'TOTAL']
            # COLUMN_ORDER.extend(AGE_OF_DEATH_ORDER)
            # save_original_data(file, '1 minus TOT (EAM)', 'EAM', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_1_MINUS_TOT_DATA_TABLE)
            # save_original_data(file, '1 minus TOT (NEAM)', 'NEAM', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_1_MINUS_TOT_DATA_TABLE)
            # save_original_data(file, '1 minus TOT (EAF)', 'EAF', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_1_MINUS_TOT_DATA_TABLE)
            # save_original_data(file, '1 minus TOT (NEAF)', 'NEAF', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_1_MINUS_TOT_DATA_TABLE)
            #
            # if(cancer_type=='All Causes'): # for All causes there is no Raw Adj sheets.. It uses Raw Data
            #
            #     save_original_data(file, 'Raw Data (EAM)', 'EAM', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #     save_original_data(file, 'Raw Data (NEAM)', 'NEAM', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #     save_original_data(file, 'Raw Data (EAF)', 'EAF', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #     save_original_data(file, 'Raw Data (NEAF)', 'NEAF', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #
            # else:
            #     save_original_data(file, 'Raw Adj (EAM)', 'EAM', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #     save_original_data(file, 'Raw Adj (NEAM)', 'NEAM', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #     save_original_data(file, 'Raw Adj (EAF)', 'EAF', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)
            #     save_original_data(file, 'Raw Adj (NEAF)', 'NEAF', cancer_type, COLUMN_ORDER[:-1], "A:AB", ORIG_RAW_ADJUSTED_MORTALITY_DATA_TABLE)


            # save_mortality_data_by_birth_year(file, 'Mortality by birth year (EAM)', 'EAM', cancer_type, "A:GZ", ORIG_MORTALITY_BY_BIRTH_YEAR_TABLE)
            # save_mortality_data_by_birth_year(file, 'Mortality by birth year (NEAM)', 'NEAM', cancer_type, "A:GZ", ORIG_MORTALITY_BY_BIRTH_YEAR_TABLE)
            # save_mortality_data_by_birth_year(file, 'Mortality by birth year (EAF)', 'EAF', cancer_type, "A:GZ", ORIG_MORTALITY_BY_BIRTH_YEAR_TABLE)
            # save_mortality_data_by_birth_year(file, 'Mortality by birth year (NEAF)', 'NEAF', cancer_type, "A:GZ", ORIG_MORTALITY_BY_BIRTH_YEAR_TABLE)





def save_original_data(file_name,sheet_name,ethnicity,cancer_type,COLUMN_ORDER,USE_COLS,table_name):

   if((cancer_type in FEMALE_ONLY_CANCER and ethnicity in['EAM','NEAM']) or
           (cancer_type  in MALE_ONLY_CANCER and ethnicity in ['EAF', 'NEAF'])):
       print("Skipping Data:", sheet_name, " for  cancer:", cancer_type, " Ethnicity:", ethnicity)
       return;

   print("Loading Data:",sheet_name," for  cancer:", cancer_type," Ethnicity:",ethnicity)
   df_ea=read_existing_data(file_name, sheet_name, COLUMN_ORDER, USE_COLS)
   if (df_ea.shape[0]>0):
       df_ea['CANCER_TYPE'] = cancer_type
       df_ea['ETHNICITY'] = ethnicity
       # df_ea.to_sql(ORIG_MORTALITY_DATA_TABLE, sqlite_connection, if_exists='append', index=False)

       #there is a issue with sum of under 1 year to 4 years and under_5_years column in the old data. so calcuate again here
       df_ea['under_5_years'] = df_ea[['under_one_year', 'year_1', 'years_2', 'years_3', 'years_4']].sum(axis=1)
       insert_original_raw_mortality_data(df_ea, ethnicity, cancer_type,table_name)


def save_mortality_data_by_birth_year(file_name,sheet_name,ethnicity,cancer_type,USE_COLS,table_name):
    if ((cancer_type in ['Female Genital Cancer-Other', 'Ovarian Cancer', 'Female Breast Cancer'] and ethnicity in ['EAM', 'NEAM']) or
            (cancer_type in ['Male Breast Cancer'] and ethnicity in ['EAF', 'NEAF'])):
        print("Skipping Data:", sheet_name, " for  cancer:", cancer_type, " Ethnicity:", ethnicity)
        return;

    print('Started Loading Data for Cancer: ', cancer_type, " Ethnicity:", ethnicity, ' Table: ', table_name)
    df_ea = df_from_excel(file_name, sheet_name, USE_COLS)
    df_ea = df_ea.rename(columns=df_ea.iloc[0]) #use first row as column headers
    df_ea=df_ea.drop(df_ea.index[0]) #remove the row which has column names
    df_ea = df_ea.apply(pd.to_numeric, errors='ignore')
    # df_ea.drop('', axis=1, inplace=True)
    df_ea=df_ea.dropna(axis=1, how='all') # drop column if all values are null
    df_ea=df_ea.head(23) # we are interested in first 23 rows
    df_ea = df_ea.dropna(how='all') # drop all rows with all nulls
    df_ea = df_ea.dropna(subset=['Year born       Æ    Age    Ø'])
    df_ea = df_ea.melt(id_vars=['Year born       Æ    Age    Ø'], value_vars=df_ea.columns.drop(['Year born       Æ    Age    Ø']).to_list(), var_name='YEAR_OF_BIRTH', value_name='OBS_BY_BIRTH_YEAR')
    df_ea['YEAR']=df_ea['YEAR_OF_BIRTH']+df_ea['Year born       Æ    Age    Ø']
    df_ea['YEAR'] = np.ceil(df_ea['YEAR']).astype(int)
    df_ea.rename(columns={ 'Year born       Æ    Age    Ø':'AGE_AT_DEATH'}, inplace=True)
    df_ea[['AGE_AT_DEATH','YEAR_OF_BIRTH','OBS_BY_BIRTH_YEAR']]
    df_ea['CANCER_TYPE'] = cancer_type
    df_ea['ETHNICITY'] = ethnicity
    insert_original_raw_mortality_data(df_ea, ethnicity, cancer_type,table_name)



def insert_original_raw_mortality_data(df_data,ethnicity,cancer_type,table_name):

    #this is to take care of ' in the cancer name. e.g. Hodgkin's cancer. SQL query doesn't like "'" in the name
    if ("'" in cancer_type):
        cancer_type = cancer_type.replace("'", "''")

    #check if table exists
    df_table = pd.read_sql_query(
        "SELECT  name  from sqlite_master where type='table' AND name='"+table_name+"'",
        sqlite_connection)

    #if it exists then delete existing data
    if(df_table.shape[0]>0):
        df = pd.read_sql_query(
            "SELECT *  from " + table_name +" WHERE cancer_type='"+cancer_type+"' and ETHNICITY='"+ethnicity+"'",
            sqlite_connection)
        if(df.shape[0]>0): # if the data is already there then remove it
            # print('Deleting existing Data for  ',cancer_type,' from ',table_name)

            sql = "DELETE FROM "+table_name+" WHERE cancer_type='"+cancer_type+"' and ETHNICITY='"+ethnicity+"'"

            sqlite_connection.execute(sql)

    #append/insert the new data
    df_data.to_sql(table_name, sqlite_connection, if_exists='append', index=False)
    print('Completed Loading Data for Cancer: ', cancer_type, " Ethnicity:",ethnicity,' Table: ', table_name)





# ==================================================================================================================================================




def generate_mit_data_files():


    #get the mapping of mit to uce codes
    df = get_mit_ucr358_mapping()
    # drop rows if there is no ucr code attached- we wont be able to do anything with this anyway
    # df = df.dropna(subset=(['UCR358_CODE','ICD10_CODE','MIT_DATA_FILE_NAME']),how='all')

    df = df.dropna(subset=(['MIT_DATA_FILE_NAME']))

    # df = df.astype({"START_YEAR": int})
    df=df.where(df.notnull(), None)

    #get the population data . This is same for each cancertype
    df_population=get_population_data()
    df_population = df_population.astype({"YEAR": int})

    df_1_tot_adjustment,df_1_tot_adjustment_wide=load_tot_adjustment_factor()
    df_1_tot_adjustment = df_1_tot_adjustment.astype({"YEAR": int})
    df_1_tot_adjustment_wide = df_1_tot_adjustment_wide.astype({"YEAR": int})






    for cancer_data_file in df.MIT_DATA_FILE_NAME.unique():

        #check if this ccancery type is flagged for generating data file
        # generate_file_flag=df[df['MIT_DATA_FILE_NAME']==cancer_data_file]['GENERATE_FILE_FLAG'].values[0]
        update_raw_data_year = df[df['MIT_DATA_FILE_NAME'] == cancer_data_file]['UPDATE_RAW_DATA_YEAR'].values[0]
        update_population_year = df[df['MIT_DATA_FILE_NAME'] == cancer_data_file]['UPDATE_POPULATION_YEAR'].values[0]
        update_tot_adj_year = df[df['MIT_DATA_FILE_NAME'] == cancer_data_file]['UPDATE_TOT_ADJ_YEAR'].values[0]
        update_1_minus_tot_year = df[df['MIT_DATA_FILE_NAME'] == cancer_data_file]['UPDATE_1_MINUS_TOT_ADJ_YEAR'].values[0]



        if(update_raw_data_year is None and update_population_year is None and update_tot_adj_year is None and update_1_minus_tot_year is None):
            print("Skipping processing for file",cancer_data_file)
            continue





        cancer_type = cancer_data_file.split('.')[0]

        # cancer_type_display=cancer_type.replace(" ", "")
        cancer_type_txt_file = cancer_type.replace('_06', '')

        cancer_type = cancer_type.replace('Cancer of the', '').replace('Cancer of', '').replace('_06', '')
        cancer_type = cancer_type.strip()
        data_file_name = os.path.join(OUTPUT_DATA_FILE_PATH, cancer_data_file)
        book = load_workbook(data_file_name)
        writer = pd.ExcelWriter(data_file_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        print("Started Processing :=================================================", cancer_type)

        try:

            if (cancer_type not in FEMALE_ONLY_CANCER):
                ss_sheet = writer.sheets['Raw Data (EAM)']
                ss_sheet.title = '(EAM) MOR(t)'
                ss_sheet = writer.sheets['Raw Data (NEAM)']
                ss_sheet.title = '(NEAM) MOR(t)'
            if (cancer_type not in MALE_ONLY_CANCER):
                ss_sheet = writer.sheets['Raw Data (EAF)']
                ss_sheet.title = '(EAF) MOR(t)'
                ss_sheet = writer.sheets['Raw Data (NEAF)']
                ss_sheet.title = '(NEAF) MOR(t)'



            # if (cancer_type not in FEMALE_ONLY_CANCER):
            #     ss_sheet = writer.sheets['(EAM) MOR(t)']
            #     ss_sheet.title = 'A_(EAM) MOR(t)'
            #
            #     ss_sheet = writer.sheets['1 minus TOT (EAM)']
            #     ss_sheet.title = 'B_1 minus TOT (EAM)'
            #
            #     ss_sheet = writer.sheets['Raw Adj (EAM)']
            #     ss_sheet.title = 'C_Raw Adj (EAM)'
            #
            #     ss_sheet = writer.sheets['OBS (EAM)']
            #     ss_sheet.title = 'D_OBS (EAM)'
            #
            #     ss_sheet = writer.sheets['Population (EAM)']
            #     ss_sheet.title = 'E_Population (EAM)'
            #
            #     ss_sheet = writer.sheets['Mortality by birth year (EAM)']
            #     ss_sheet.title = 'F_OBS star (EAM)'
            #
            #
            #     ss_sheet = writer.sheets['(NEAM) MOR(t)']
            #     ss_sheet.title = 'M_(NEAM) MOR(t)'
            #
            #     ss_sheet = writer.sheets['1 minus TOT (NEAM)']
            #     ss_sheet.title = 'N_1 minus TOT (NEAM)'
            #
            #     ss_sheet = writer.sheets['Raw Adj (NEAM)']
            #     ss_sheet.title = 'O_Raw Adj (NEAM)'
            #
            #     ss_sheet = writer.sheets['OBS (NEAM)']
            #     ss_sheet.title = 'P_OBS (NEAM)'
            #
            #     ss_sheet = writer.sheets['Population (NEAM)']
            #     ss_sheet.title = 'Q_Population (NEAM)'
            #
            #     ss_sheet = writer.sheets['Mortality by birth year (NEAM)']
            #     ss_sheet.title = 'R_OBS star (NEAM)'
            #
            #
            # if (cancer_type not in MALE_ONLY_CANCER):
            #     ss_sheet = writer.sheets['(EAF) MOR(t)']
            #     ss_sheet.title = 'G_(EAF) MOR(t)'
            #
            #     ss_sheet = writer.sheets['1 minus TOT (EAF)']
            #     ss_sheet.title = 'H_1 minus TOT (EAF)'
            #
            #     ss_sheet = writer.sheets['Raw Adj (EAF)']
            #     ss_sheet.title = 'I_Raw Adj (EAF)'
            #
            #     ss_sheet = writer.sheets['OBS (EAF)']
            #     ss_sheet.title = 'J_OBS (EAF)'
            #
            #     ss_sheet = writer.sheets['Population (EAF)']
            #     ss_sheet.title = 'K_Population (EAF)'
            #
            #     ss_sheet = writer.sheets['Mortality by birth year (EAF)']
            #     ss_sheet.title = 'L_OBS star (EAF)'
            #
            #     ss_sheet = writer.sheets['(NEAF) MOR(t)']
            #     ss_sheet.title = 'S_(NEAF) MOR(t)'
            #
            #     ss_sheet = writer.sheets['1 minus TOT (NEAF)']
            #     ss_sheet.title = 'T_1 minus TOT (NEAF)'
            #
            #     ss_sheet = writer.sheets['Raw Adj (NEAF)']
            #     ss_sheet.title = 'U_Raw Adj (NEAF)'
            #
            #     ss_sheet = writer.sheets['OBS (NEAF)']
            #     ss_sheet.title = 'V_OBS (NEAF)'
            #
            #     ss_sheet = writer.sheets['Population (NEAF)']
            #     ss_sheet.title = 'W_Population (NEAF)'
            #
            #     ss_sheet = writer.sheets['Mortality by birth year (NEAF)']
            #     ss_sheet.title = 'X_OBS star (NEAF)'
            #
            # ss_sheet = writer.sheets['Decades (EA)']
            # ss_sheet.title = 'Y_Decades (EA)'
            #
            # ss_sheet = writer.sheets['Decades (NEA)']
            # ss_sheet.title = 'Z_Decades (NEA)'
            #
            #
            # writer.book._sheets.sort(key=lambda ws: ws.title)
            # print("Renaming is done")
            #
            # for i in writer.book.worksheets:
            #     if i.title in [ '0.5-3 (EA)','0.5-3 (NEA)','22.5-42.5 (EA)','22.5-42.5 (NEA)','52.5-72.5 (EA)','52.5-72.5 (NEA)','7.5-17.5 (EA)','7.5-17.5 (NEA)','82.5-102.5 (EA)','82.5-102.5 (NEA)','Early(EA)','Early (NEA)','Middle (EA)','Middle (NEA)','Mortality Chart (EA)','Mortality Chart (NEA)','Late (EA)','Late (NEA)','Early (EA)']:
            #         writer.book.remove_sheet(i)
            #
            #
            # print("removed sheets")
            #
            #
            # # #rename it back
            # # for i in writer.book.worksheets:
            # #
            # # a_string[2:]
            # writer.book.active = 0
            #

            # #rename it back
            # for i in writer.book.worksheets:
            #     ss_sheet = writer.sheets[i.title]
            #     ss_sheet.title = i.title[2:]


            writer.save()
            writer.close()

        except Exception:
            print("Exception=================",cancer_type)
            writer.save()
            writer.close()
            # continue

        writer = pd.ExcelWriter(data_file_name, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        PATH = (base_path / "ucr_output_data_files/").resolve()
        load_existing_mortality_raw_data(PATH,cancer_type)

        # get Raw data start years
        df_mortality_start_years = get_raw_mortality_start_years()
        df_mortality_start_years = df_mortality_start_years.astype({"START_YEAR": int})




        if (cancer_type not in FEMALE_ONLY_CANCER):
            eam_mortality_start_year = df_mortality_start_years[(df_mortality_start_years['CANCER_TYPE'] == cancer_type) & (df_mortality_start_years['ETHNICITY'] == 'EAM')]['START_YEAR'].values[0]
            neam_mortality_start_year = df_mortality_start_years[(df_mortality_start_years['CANCER_TYPE'] == cancer_type) & (df_mortality_start_years['ETHNICITY'] == 'NEAM')]['START_YEAR'].values[0]
        if (cancer_type not in MALE_ONLY_CANCER):
            eaf_mortality_start_year = df_mortality_start_years[(df_mortality_start_years['CANCER_TYPE'] == cancer_type) & (df_mortality_start_years['ETHNICITY'] == 'EAF')]['START_YEAR'].values[0]
            neaf_mortality_start_year = df_mortality_start_years[(df_mortality_start_years['CANCER_TYPE'] == cancer_type) & (df_mortality_start_years['ETHNICITY'] == 'NEAF')]['START_YEAR'].values[0]


        print("Processing Cancer:",cancer_type)

        if(update_raw_data_year is not None):

            print(" Started Processing Cancer Type:",cancer_type)
            #get the year we have to start generating data
            start_year = df[df['MIT_DATA_FILE_NAME'] == cancer_data_file]['UPDATE_RAW_DATA_YEAR'].values[0]

            #get the UCR codes for this data file
            ucr358_list=df[df['MIT_DATA_FILE_NAME']==cancer_data_file]['UCR358_CODE'].values[0]

            if(ucr358_list is None): #check if it hsd ICD10 code
                ICD10_list = df[df['MIT_DATA_FILE_NAME'] == cancer_data_file]['ICD10_CODE'].values[0]
                ICD10_list = ICD10_list.split(",")

                ICD10_list=', '.join(f'"{w}"' for w in ICD10_list)
                df_raw_data = get_ucr358_raw_data(ICD10_list,'ICD10_CODE')

            else:
                df_raw_data=get_ucr358_raw_data(ucr358_list,'UCR358_CODE')

            df_raw_data = df_raw_data.astype({"YEAR": int})

            mit_diseases_category=df[df['MIT_DATA_FILE_NAME']==cancer_data_file]['MIT_DISEASES_CATEGORY'].values[0]
            df_raw_data['MIT_DISEASES_CATEGORY']=mit_diseases_category

            df_raw_data['CANCER_TYPE']=cancer_type

            #save this data in to a sqlite table
            # df_raw_data.to_sql(MIT_RAW_MORTALITY_DATA_TABLE, sqlite_connection, if_exists='append', index=False)
            insert_mit_grouped_data(df_raw_data, cancer_type)

            # Update the existing sheets with the new data as per stated start year e.g. 2010 onwards
            df_raw_data=df_raw_data[df_raw_data['YEAR']>=start_year]

            print("Processing File:",data_file_name," starting year:",start_year)
            COLUMN_ORDER = ['YEAR', 'TOTAL']
            COLUMN_ORDER.extend(AGE_OF_DEATH_ORDER)


            print(df_raw_data.head())
            if (cancer_type not in FEMALE_ONLY_CANCER):
                df_eam = df_raw_data[df_raw_data['ETHNICITY'] == 'EAM']
                write_data_excel(writer, df_eam, '(EAM) MOR(t)', COLUMN_ORDER, "A:AC", update_raw_data_year, data_file_name)


                df_eam = df_raw_data[df_raw_data['ETHNICITY'] == 'NEAM']
                write_data_excel(writer, df_eam, '(NEAM) MOR(t)', COLUMN_ORDER, "A:AC", update_raw_data_year, data_file_name)





            if (cancer_type not in MALE_ONLY_CANCER):  # EAF/NEAF sheets are not there or not updated for these cancer types , so skip them

                df_eam = df_raw_data[df_raw_data['ETHNICITY'] == 'EAF']
                write_data_excel(writer, df_eam, '(EAF) MOR(t)', COLUMN_ORDER, "A:AC", update_raw_data_year, data_file_name)

                df_eam = df_raw_data[df_raw_data['ETHNICITY'] == 'NEAF']
                write_data_excel(writer, df_eam, '(NEAF) MOR(t)', COLUMN_ORDER, "A:AC", update_raw_data_year, data_file_name)


            print(" Completed Processing Cancer Type:", cancer_type)
        #====Update population Data
        if(update_population_year is not None):
            df_cancer_population = df_population[df_population['YEAR'] >= update_population_year]
            df_cancer_population=df_cancer_population.apply(pd.to_numeric, errors='ignore')


            COLUMN_ORDER = ['YEAR', 'ALL_AGES', 'under_one_year', 'years_1_4']
            COLUMN_ORDER.extend(AGE_OF_DEATH_POPULATION_COLS)

            df_cancer_population[COLUMN_ORDER] = df_cancer_population[COLUMN_ORDER].astype(int)
            SUM_COLS = [age for age in COLUMN_ORDER if age not in ['YEAR', 'ETHNICITY', 'Not_stated','ALL_AGES']]
            df_cancer_population['ALL_AGES'] = df_cancer_population[SUM_COLS].sum(axis=1)


            if (cancer_type not in FEMALE_ONLY_CANCER):
                df_eam = df_cancer_population[df_cancer_population['ETHNICITY'] == 'EAM']

                df_eam=df_eam[df_eam['YEAR']>=eam_mortality_start_year]
                write_data_excel(writer, df_eam, 'Population (EAM)', COLUMN_ORDER,"A:AB",update_population_year,data_file_name)
                df_eam = df_cancer_population[df_cancer_population['ETHNICITY'] == 'NEAM']
                df_eam = df_eam[df_eam['YEAR'] >= neam_mortality_start_year]
                write_data_excel(writer, df_eam, 'Population (NEAM)', COLUMN_ORDER,"A:AB",update_population_year,data_file_name)

            if (cancer_type not in MALE_ONLY_CANCER):  # EAF/NEAF sheets are not there or not updated for these cancer types , so skip them

                df_eam = df_cancer_population[df_cancer_population['ETHNICITY'] == 'EAF']
                df_eam = df_eam[df_eam['YEAR'] >= eaf_mortality_start_year]
                write_data_excel(writer, df_eam, 'Population (EAF)', COLUMN_ORDER,"A:AB",update_population_year,data_file_name)
                df_eam = df_cancer_population[df_cancer_population['ETHNICITY'] == 'NEAF']
                df_eam = df_eam[df_eam['YEAR'] >= neaf_mortality_start_year]
                write_data_excel(writer, df_eam, 'Population (NEAF)', COLUMN_ORDER,"A:AB",update_population_year,data_file_name)

        print("Completed Population data")
        # ====Update 1-ToT sheets
        if (update_1_minus_tot_year is not None):

            df_1_tot_adjustment = df_1_tot_adjustment[(df_1_tot_adjustment['YEAR'] >= update_tot_adj_year)]
            df_1_tot_adjustment = df_1_tot_adjustment.apply(pd.to_numeric, errors='ignore')

            df_1_tot_adjustment_wide = df_1_tot_adjustment_wide[(df_1_tot_adjustment_wide['YEAR'] >= update_tot_adj_year)]
            df_1_tot_adjustment_wide = df_1_tot_adjustment_wide.apply(pd.to_numeric, errors='ignore')



            COLUMN_ORDER = ['YEAR', 'TOTAL']
            COLUMN_ORDER.extend(AGE_OF_DEATH_ORDER)
            if (cancer_type not in FEMALE_ONLY_CANCER):
                df_eam = df_1_tot_adjustment_wide[df_1_tot_adjustment_wide['ETHNICITY'] == 'EAM']
                df_eam = df_eam[df_eam['YEAR'] >= eam_mortality_start_year]
                write_data_excel(writer, df_eam, '1 minus TOT (EAM)', COLUMN_ORDER[:-1],"A:AB",update_tot_adj_year,data_file_name)
                df_eam = df_1_tot_adjustment_wide[df_1_tot_adjustment_wide['ETHNICITY'] == 'NEAM']
                df_eam = df_eam[df_eam['YEAR'] >= neam_mortality_start_year]
                write_data_excel(writer, df_eam, '1 minus TOT (NEAM)', COLUMN_ORDER[:-1],"A:AB",update_tot_adj_year,data_file_name)

            if (cancer_type not in MALE_ONLY_CANCER):  # EAF/NEAF sheets are not there or not updated for these cancer types , so skip them

                df_eam = df_1_tot_adjustment_wide[df_1_tot_adjustment_wide['ETHNICITY'] == 'EAF']
                df_eam = df_eam[df_eam['YEAR'] >= eaf_mortality_start_year]
                write_data_excel(writer, df_eam, '1 minus TOT (EAF)', COLUMN_ORDER[:-1],"A:AB",update_tot_adj_year,data_file_name)
                df_eam = df_1_tot_adjustment_wide[df_1_tot_adjustment_wide['ETHNICITY'] == 'NEAF']
                df_eam = df_eam[df_eam['YEAR'] >= neaf_mortality_start_year]
                write_data_excel(writer, df_eam, '1 minus TOT (NEAF)', COLUMN_ORDER[:-1],"A:AB",update_tot_adj_year,data_file_name)

            print("Completed 1-ToT adjustment data")

            # ====Update TOT Adjustment
            # if (update_tot_adj_year is not None):
            if (True):
                COLUMN_ORDER = ['YEAR', 'TOTAL']
                COLUMN_ORDER.extend(AGE_OF_DEATH_ORDER)



                # get the raw mortlaity data
                if (cancer_type not in FEMALE_ONLY_CANCER):  # EAM/NEAM sheets are not present in the files for these two cancer types

                    df_tot, df_tot_adjustment_wide,df_OBS_wide = load_raw_adjustment_mortality(cancer_type, df_1_tot_adjustment)
                    df_decades_mortality = load_birth_year_cohort_mortality(cancer_type,df_tot.copy(),10)
                    df_decades_mortality=df_decades_mortality[df_decades_mortality.columns & COL_ORDER_DECADES]

                    df_yearly_mortality = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 1)

                    #this is required for TXT file generation
                    df_10_year = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 10, True)
                    df_5_year = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 5, True)
                    df_1_year = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 1, True)
                    generate_txt_files(df_10_year,cancer_type, 10, 'EAM',cancer_type_txt_file)
                    generate_txt_files(df_5_year,cancer_type, 5, 'EAM',cancer_type_txt_file)
                    generate_txt_files(df_10_year,cancer_type, 10, 'NEAM',cancer_type_txt_file)
                    generate_txt_files(df_5_year,cancer_type, 5, 'NEAM',cancer_type_txt_file)

                    df_tot_eam = df_tot_adjustment_wide[df_tot_adjustment_wide['ETHNICITY'] == 'EAM']
                    write_data_excel(writer, df_tot_eam, 'Raw Adj (EAM)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)

                    df_OBS_eam = df_OBS_wide[df_OBS_wide['ETHNICITY'] == 'EAM']
                    write_data_excel(writer, df_OBS_eam, 'OBS (EAM)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)


                    df_tot_eam = df_tot_adjustment_wide[df_tot_adjustment_wide['ETHNICITY'] == 'NEAM']
                    write_data_excel(writer, df_tot_eam, 'Raw Adj (NEAM)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)

                    df_OBS_eam = df_OBS_wide[df_OBS_wide['ETHNICITY'] == 'NEAM']
                    write_data_excel(writer, df_OBS_eam, 'OBS (NEAM)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)

                    df_decades_mortality_eam=df_decades_mortality[df_decades_mortality['ETHNICITY'] == 'EAM']
                    write_decades_data_excel(writer, df_decades_mortality_eam, "EAM", "Decades (EA)",cancer_type)

                    df_decades_mortality_neam = df_decades_mortality[df_decades_mortality['ETHNICITY'] == 'NEAM']
                    write_decades_data_excel(writer, df_decades_mortality_neam, "NEAM", "Decades (NEA)",cancer_type)

                    df_yearly_mortality_eam=df_yearly_mortality[df_yearly_mortality['ETHNICITY'] == 'EAM']
                    write_yearly_data_excel(writer, df_yearly_mortality_eam, "EAM", "Mortality by birth year (EAM)",cancer_type)

                    df_yearly_mortality_neam = df_yearly_mortality[df_yearly_mortality['ETHNICITY'] == 'NEAM']
                    write_yearly_data_excel(writer, df_yearly_mortality_neam, "NEAM", "Mortality by birth year (NEAM)",cancer_type)






                if (cancer_type not in MALE_ONLY_CANCER):  # EAF/NEAF sheets are not there or not updated for these cancer types , so skip them

                    df_tot, df_tot_adjustment_wide,df_OBS_wide = load_raw_adjustment_mortality(cancer_type, df_1_tot_adjustment)
                    df_decades_mortality = load_birth_year_cohort_mortality(cancer_type,df_tot.copy(), 10)
                    df_decades_mortality = df_decades_mortality[df_decades_mortality.columns & COL_ORDER_DECADES]

                    df_yearly_mortality = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 1)

                    #this is required for TXT file generation
                    df_10_year = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 10, True)
                    df_5_year = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 5, True)
                    df_1_year = load_birth_year_cohort_mortality(cancer_type, df_tot.copy(), 1, True)




                    # df_tot_eam = get_tot_adjustment(data_file_name,cancer_type,'Raw Data (EAF)', '1 minus TOT (EAF)', 'EAF')
                    df_tot_eam = df_tot_adjustment_wide[df_tot_adjustment_wide['ETHNICITY'] == 'EAF']
                    write_data_excel(writer, df_tot_eam, 'Raw Adj (EAF)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)

                    df_OBS_eam = df_OBS_wide[df_OBS_wide['ETHNICITY'] == 'EAF']
                    write_data_excel(writer, df_OBS_eam, 'OBS (EAF)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)



                    df_tot_eam = df_tot_adjustment_wide[df_tot_adjustment_wide['ETHNICITY'] == 'NEAF']
                    write_data_excel(writer, df_tot_eam, 'Raw Adj (NEAF)', COLUMN_ORDER[:-1],"A:AB", update_tot_adj_year, data_file_name)

                    df_OBS_eam = df_OBS_wide[df_OBS_wide['ETHNICITY'] == 'NEAF']
                    write_data_excel(writer, df_OBS_eam, 'OBS (NEAF)', COLUMN_ORDER[:-1], "A:AB", update_tot_adj_year, data_file_name)


                    df_decades_mortality_eaf=df_decades_mortality[df_decades_mortality['ETHNICITY'] == 'EAF']
                    write_decades_data_excel(writer, df_decades_mortality_eaf, "EAF", "Decades (EA)",cancer_type)

                    df_decades_mortality_neaf = df_decades_mortality[df_decades_mortality['ETHNICITY'] == 'NEAF']
                    write_decades_data_excel(writer, df_decades_mortality_neaf, "NEAF", "Decades (NEA)",cancer_type)

                    print("Completed Writing Decades Sheet")

                    df_yearly_mortality_eam=df_yearly_mortality[df_yearly_mortality['ETHNICITY'] == 'EAF']
                    write_yearly_data_excel(writer, df_yearly_mortality_eam, "EAF", "Mortality by birth year (EAF)",cancer_type)

                    df_yearly_mortality_neam = df_yearly_mortality[df_yearly_mortality['ETHNICITY'] == 'NEAF']
                    write_yearly_data_excel(writer, df_yearly_mortality_neam, "NEAF", "Mortality by birth year (NEAF)",cancer_type)

                    #create TXT Files
                    generate_txt_files(df_10_year,cancer_type, 10, 'EAF',cancer_type_txt_file)
                    generate_txt_files(df_5_year,cancer_type, 5, 'EAF',cancer_type_txt_file)
                    generate_txt_files(df_10_year,cancer_type, 10, 'NEAF',cancer_type_txt_file)
                    generate_txt_files(df_5_year,cancer_type, 5, 'NEAF',cancer_type_txt_file)




        writer.save()
        writer.close()
        print("Completed Processing :=================================================", cancer_type)



def insert_mit_grouped_data(df_data,cancer_type):
    df = pd.read_sql_query(
        "SELECT distinct CANCER_TYPE  from " + MIT_RAW_MORTALITY_DATA_TABLE,
        sqlite_connection)
    if(df.shape[0]>0): # if the dat is already there then remove it
        print('Deleting existing Data for cancer type',cancer_type,' from ',MIT_RAW_MORTALITY_DATA_TABLE)
        if ("'" in cancer_type):
               cancer_type = cancer_type.replace("'", "''")


        sql = "DELETE FROM "+MIT_RAW_MORTALITY_DATA_TABLE+" WHERE CANCER_TYPE='"+cancer_type+"'"

        sqlite_connection.execute(sql)

    #append the new data
    df_data.to_sql(MIT_RAW_MORTALITY_DATA_TABLE, sqlite_connection, if_exists='append', index=False)

import xlwings as xl
def df_from_excel(path,sheet,USE_COLS):
    app = xl.App(visible=False)
    book = app.books.open(path)
    book.save()
    app.kill()
    return pd.read_excel(path,sheet_name=sheet,usecols=USE_COLS, header=None)
def read_existing_data(file_name,sheet,COLUMN_ORDER,USE_COLS,start_year=None):
    # reader = pd.ExcelFile(file_name)
    #read the existing data
    # df_ea = reader.parse(sheet, usecols=USE_COLS, header=None)
    try:
        df_ea = df_from_excel(file_name,sheet,USE_COLS)
    except Exception:
        print(" Error reading sheet:",file_name,sheet)
        df_ea=pd.DataFrame()


    if(df_ea.shape[0]>0):

        if(COLUMN_ORDER is not None):
            df_ea.columns = COLUMN_ORDER
        else:
            df_ea = df_ea.rename(columns=df_ea.iloc[0])  # use first row as
            df_ea = df_ea.drop(df_ea.index[0])  # remove the row which has column names


        df_ea = df_ea[pd.to_numeric(df_ea['YEAR'], errors='coerce').notnull()]
        df_ea = df_ea.apply(pd.to_numeric, errors='ignore')
        df_ea=df_ea[~df_ea['YEAR'].isin(['Age       Æ    Year    Ø','Mortality by all Forms of Death','Mortality by Accidents (includes undetermined)'])]
        df_ea = df_ea.astype({"YEAR": int})

        #remove the existing data from the dataframe
        if(start_year is not None):
            df_ea=df_ea[df_ea['YEAR']<start_year]
    return df_ea



def write_data_excel(writer,df,sheet,COLUMN_ORDER,USECOLS=None,start_year=None,data_file_name=None):
    print("   Processing Data:", sheet)

    df_existing = read_existing_data(data_file_name, sheet, COLUMN_ORDER, USECOLS, start_year)
    df.drop('ETHNICITY', axis=1, inplace=True)
    #update only the data which has been asked to update
    df=df[COLUMN_ORDER]
    df=df[df['YEAR']>=start_year]
    df=pd.concat([df_existing,df])
    df[COLUMN_ORDER] = df[COLUMN_ORDER].round(4)

    # there is a issue with Mortality sum of under 1 year to 4 years and under_5_years column in the old data. so calcuate again here
    if sheet.find('MOR') != -1:
     df['under_5_years'] = df[['under_one_year', 'year_1', 'years_2', 'years_3', 'years_4']].sum(axis=1)

    #delete existing data and then write new data set
    #check if this sheet exists
    try:
        mortality_sheet=writer.sheets[sheet]
        mortality_sheet.delete_rows(2, mortality_sheet.max_row - 1)
        df.to_excel(writer, sheet, index=False, header=False, startrow=1, startcol=0)
    except:
        # mortality_sheet=writer.book.create_sheet(sheet)
        df.to_excel(writer, sheet, index=False, header=True, startrow=0, startcol=0)
        print(" Created a new sheet :",sheet)

    return df


def write_decades_data_excel(writer,df,ethnicity,sheet,cancer_type):

    print("   Processing Data:", sheet)
    df.drop('ETHNICITY', axis=1, inplace=True)

    sht = writer.sheets[sheet]

    if(ethnicity in ['EAM','NEAM']):
        # writer.sheets[sheet].delete_rows(4, 27)
        start_row=3
        sht['J1'] = "(" + ethnicity + ") " + "OBS*(h,t)"
        sht['I2'] = ""
        sht['I1'] = ""
        sht['J2'] = ""

    else:
        start_row=30
        sht['J28'] = "(" + ethnicity + ") " + "OBS*(h,t)"
        sht['I29'] = ""
        sht['I28'] = ""
        sht['J29'] = ""
        # writer.sheets[sheet].delete_rows(32, 54)


    df.to_excel(writer, sheet, index=False, header=True, startrow=start_row, startcol=0)

    BODY_FONT = Font(name='Times New Roman',
                         size=36,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='000000')


    for rownum in range(4, 26):
        for colnum in range (1,30):
            sht.cell(row=rownum ,column=colnum).font = BODY_FONT

    for rownum in range(31, 53):
        for colnum in range (1,30):
            sht.cell(row=rownum ,column=colnum).font = BODY_FONT

    set_border(sht, 'A4:W26')
    set_border(sht, 'A31:W53')




def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)



def write_yearly_data_excel(writer,df,ethnicity,sheet,cancer_type):

    print("   Processing Data:", sheet)
    df.drop('ETHNICITY', axis=1, inplace=True)
    df = df.dropna(axis=1, how='all')

    # if(ethnicity in ['EAM','NEAM']):
    #     # writer.sheets[sheet].delete_rows(4, 27)
    #     start_row=3
    # else:
    #     start_row=30
    #     # writer.sheets[sheet].delete_rows(32, 54)


    df.to_excel(writer, sheet, index=False, header=True, startrow=0, startcol=0)


def create_old_new_comparison_sheet():

    #get the newly generated data
    df = pd.read_sql_query(
        "SELECT CANCER_TYPE,YEAR,ETHNICITY,TOTAL as NEW_PROCESSED_TOTAL  from " + MIT_RAW_MORTALITY_DATA_TABLE +" order by CANCER_TYPE,YEAR,ETHNICITY",
        sqlite_connection)

    #get the old/existing data
    df_orig = pd.read_sql_query(
         "SELECT CANCER_TYPE,YEAR,ETHNICITY,TOTAL  from " + ORIG_MORTALITY_DATA_TABLE +" order by CANCER_TYPE,YEAR,ETHNICITY",
        sqlite_connection)

    df_merged=pd.merge(df,df_orig,how='left',on=['CANCER_TYPE','YEAR','ETHNICITY'])
    df_merged['DIFF']=df_merged['NEW_PROCESSED_TOTAL']-df_merged['EXISTING_TOTAL']

    df_merged.loc[df_merged['EXISTING_TOTAL']<0,'DIFF']=''

    filename=data_file_name = os.path.join(OUTPUT_DATA_FILE_PATH, "mortality_data_comparison.xlsx")
    df_merged.to_excel(filename,index=False)









#==============create Mortality Data Rectangle===================================

# def create_mit_data_rectangle(DATA_FILE_PATH):
#     # DATA_FILE_PATH = (base_path / "mortality_orig_files/").resolve()
#     # read all files and append the data frame
#     directory = os.fsencode(DATA_FILE_PATH)
#
#     for file in os.listdir(directory):
#         filename = os.fsdecode(file)
#         if filename.endswith(".xls") or filename.endswith(".xlsx"):
#             file = os.path.join(DATA_FILE_PATH, filename)
#             cancer_xl = pd.ExcelFile(file)
#             cancer_type = filename.split('.')[0]
#             cancer_type = cancer_type.replace('Cancer of the', '').replace('Cancer of', '').replace('_06', '')
#             cancer_type = cancer_type.strip()
#             print("Loading Data for this cancer:", cancer_type)
#
#             COLUMN_ORDER_RAW_MORTALITY = ['YEAR', 'TOTAL_RAW_MORTALITY']
#             AGE_OF_DEATH_ORDER_RAW_MORTALITY=[s + '_RAW_MORTALITY' for s in AGE_OF_DEATH_ORDER]
#             COLUMN_ORDER_RAW_MORTALITY.extend(AGE_OF_DEATH_ORDER_RAW_MORTALITY)
#
#             COLUMN_ORDER_TOT_ADJUSTMENT = []
#             AGE_OF_DEATH_ORDER_TOT_ADJUSTMENT=[s + '_TOT_ADJUSTMENT' for s in AGE_OF_DEATH_ORDER]
#             COLUMN_ORDER_TOT_ADJUSTMENT.extend(AGE_OF_DEATH_ORDER_TOT_ADJUSTMENT)
#
#             COLUMN_ORDER_1_MINUS_TOT = ['YEAR', 'TOTAL_1_MINUS_TOT']
#             AGE_OF_DEATH_ORDER_1_MINUS_TOT=[s + '_1_MINUS_TOT' for s in AGE_OF_DEATH_ORDER]
#             COLUMN_ORDER_1_MINUS_TOT.extend(AGE_OF_DEATH_ORDER_1_MINUS_TOT)
#
#
#             USECOLS="A:AC"
#             # get the raw mortlaity data
#             if(cancer_type not in ['Female Genital Cancer-Other','Ovarian Cancer','Female Breast Cancer']): # EAM/NEAM sheets are not present in the files for these two cancer types
#                 # df_raw_mortality_data = read_existing_data(file, 'Raw Data (EAM)', COLUMN_ORDER_RAW_MORTALITY, USECOLS)
#                 # df_raw_mortality_data['CANCER_TYPE'] = cancer_type
#                 # df_raw_mortality_data['ETHNICITY'] = 'EAM'
#                 #
#                 # df_raw_tot_adjustment = read_existing_data(file, '1 minus TOT (EAM)', COLUMN_ORDER_1_MINUS_TOT, USECOLS)
#                 # df_raw_tot_adjustment['CANCER_TYPE'] = cancer_type
#                 # df_raw_tot_adjustment['ETHNICITY'] = 'EAM'
#                 #
#                 # #merge these two dfs
#                 # df_raw_tot=pd.merge(df_raw_mortality_data,df_raw_tot_adjustment,on=['CANCER_TYPE','YEAR','ETHNICITY'],how='left')
#                 # for raw_adj_col in AGE_OF_DEATH_ORDER:
#                 #     df_raw_tot[raw_adj_col+'_TOT_ADJUSTMENT']=df_raw_tot[raw_adj_col+'_RAW_MORTALITY']/df_raw_tot[raw_adj_col+'_1_MINUS_TOT']
#                 df_tot_eam=get_tot_adjustment('Raw Data (EAM)','1 minus TOT (EAM)','EAM')
#
#                 print(df_tot_eam.head())
#
#
#             if (cancer_type not in ['Male Breast Cancer']): # EAF/NEAF sheets are not there or not updated for these cancer types , so skip them
#                 save_original_data(cancer_xl, 'Raw Data (EAF)', 'EAF', cancer_type)
#                 save_original_data(cancer_xl, 'Raw Data (NEAF)', 'NEAF', cancer_type)
#
#
#             def get_tot_adjustment(sheet_mortality,sheet_tot,ethnicity):
#                 df_raw_mortality_data = read_existing_data(file, sheet_mortality, COLUMN_ORDER_RAW_MORTALITY, USECOLS)
#                 df_raw_mortality_data['CANCER_TYPE'] = cancer_type
#                 df_raw_mortality_data['ETHNICITY'] = ethnicity
#
#                 df_raw_tot_adjustment = read_existing_data(file, sheet_tot, COLUMN_ORDER_1_MINUS_TOT, USECOLS)
#                 df_raw_tot_adjustment['CANCER_TYPE'] = cancer_type
#                 df_raw_tot_adjustment['ETHNICITY'] = ethnicity
#
#                 #merge these two dfs
#                 df_raw_tot=pd.merge(df_raw_mortality_data,df_raw_tot_adjustment,on=['CANCER_TYPE','YEAR','ETHNICITY'],how='left')
#                 for raw_adj_col in AGE_OF_DEATH_ORDER:
#                     df_raw_tot[raw_adj_col+'_TOT_ADJUSTMENT']=df_raw_tot[raw_adj_col+'_RAW_MORTALITY']/df_raw_tot[raw_adj_col+'_1_MINUS_TOT']
#
#                 return df_raw_tot

#===========================================================================================================



#=============These methods are related to dealing with population Data===================
def load_master_population_data():

    #This is a master Population file used to load data for all cancer types
    file=os.path.join(POPULATION_FILES, "MIT_MASTER_POPULATION_DATA.xlsx")
    cancer_xl = pd.ExcelFile(file)

    COLUMN_ORDER = ['YEAR', 'ALL_AGES','under_one_year','years_1_4']
    COLUMN_ORDER.extend(AGE_OF_DEATH_POPULATION_COLS)
    save_population_data_to_db(cancer_xl, 'Population (EAM)', 'EAM',POPULATION_DATA_TABLE,COLUMN_ORDER)
    save_population_data_to_db(cancer_xl, 'Population (NEAM)', 'NEAM',POPULATION_DATA_TABLE,COLUMN_ORDER)
    save_population_data_to_db(cancer_xl, 'Population (EAF)', 'EAF',POPULATION_DATA_TABLE,COLUMN_ORDER)
    save_population_data_to_db(cancer_xl, 'Population (NEAF)', 'NEAF',POPULATION_DATA_TABLE,COLUMN_ORDER)

    print("Completed Loading Population")

    #save 1-TOT(h,t)
    COLUMN_ORDER = ['YEAR', 'TOTAL']
    COLUMN_ORDER.extend(AGE_OF_DEATH_ORDER)
    save_population_data_to_db(cancer_xl, '1 minus TOT (EAM)', 'EAM', POPULATION_TOT_ADJ_DATA_TABLE,COLUMN_ORDER[:-1])  #  take all except last (-1) because we do not want last name in the list 'Not Stated'
    save_population_data_to_db(cancer_xl, '1 minus TOT (NEAM)', 'NEAM', POPULATION_TOT_ADJ_DATA_TABLE,COLUMN_ORDER[:-1])
    save_population_data_to_db(cancer_xl, '1 minus TOT (EAF)', 'EAF', POPULATION_TOT_ADJ_DATA_TABLE,COLUMN_ORDER[:-1])
    save_population_data_to_db(cancer_xl, '1 minus TOT (NEAF)', 'NEAF', POPULATION_TOT_ADJ_DATA_TABLE,COLUMN_ORDER[:-1])



    print("Completed Loading ToT")



def save_population_data_to_db(cancer_xl,sheet_name,ethnicity,table_name,COLUMN_ORDER):

    df_ea = cancer_xl.parse(sheet_name, usecols="A:AB", header=None)
    df_ea=df_ea.dropna(how='all')

    if (df_ea.shape[0]>0):
        df_ea.columns = COLUMN_ORDER
        df_ea = df_ea.dropna(subset=['YEAR'])
        df_ea = df_ea[pd.to_numeric(df_ea['YEAR'], errors='coerce').notnull()]
        df_ea = df_ea.astype({"YEAR": int})
        df_ea['ETHNICITY'] = ethnicity

        # check if table exists
        df_table = pd.read_sql_query(
            "SELECT  name  from sqlite_master where type='table' AND name='" + table_name + "'",
            sqlite_connection)

        # if it exists then delete existing data
        if (df_table.shape[0] > 0):
                sql = "DELETE FROM " + table_name +" WHERE ETHNICITY='"+ethnicity+"'"
                sqlite_connection.execute(sql)

        # append/insert the new data
        df_ea.to_sql(table_name, sqlite_connection, if_exists='append', index=False)




def get_population_data():
    df = pd.read_sql_query(
        "SELECT *  from " + POPULATION_DATA_TABLE+" order by YEAR,ETHNICITY",
        sqlite_connection)
    return df;

def get_tot_adjustment_data():
    df = pd.read_sql_query(
        "SELECT *  from " + POPULATION_TOT_ADJ_DATA_TABLE+" order by YEAR,ETHNICITY",
        sqlite_connection)
    return df;

def get_raw_mortality_data(cancer_type=None):

    if ("'" in cancer_type):
        cancer_type = cancer_type.replace("'", "''")

    df_orig = pd.read_sql_query(
         "SELECT * from " + ORIG_MORTALITY_DATA_TABLE +" WHERE CANCER_TYPE='"+cancer_type+"' order by CANCER_TYPE,YEAR,ETHNICITY",
        sqlite_connection)
    return df_orig


def get_raw_mortality_start_years():

    df_orig = pd.read_sql_query(
         "SELECT CANCER_TYPE,ETHNICITY,min(Year) as START_YEAR from  ORIGINAL_MIT_RAW_CANCER_MORTALITY_DATA group by CANCER_TYPE,ETHNICITY order by CANCER_TYPE,YEAR,ETHNICITY",
        sqlite_connection)
    return df_orig


#=========== Population methods==========================================================================


def load_85_plus_population_data():
    """
    This functions loads the CDC population data file over 85+
    This is the file for 2019:  pcen_v2019_85to100.txt
    """

    d_file_name = os.path.join(RAW_CDC_POPULATION_FILES, "pcen_v2019_85to100.txt")
    d_lines = open(d_file_name, 'r').readlines()
    data=[]
    for line in d_lines:
        data.append({'series_vintage': line[0:4],
               'year':line[4:8],
               'estimate_month':line[8:9],
               'age':line[9:12],
               'race_sex':line[12:13],
               'hispanic_origin':line[13:14],
               'population':line[14:22]
               })

    df = pd.DataFrame(data)
    df = df.apply(pd.to_numeric, errors='ignore')
    df['ETHNICITY'] = df['race_sex'].map(cdc_85_plus_race_sex_dict)
    df['AGE_COHORT']=np.where( (df['age']>=85) & (df['age']<90),'years_85_89',
                       np.where((df['age']>=90) & (df['age']<95),'years_90_94',
                         np.where((df['age']>=95) & (df['age']<100),'years_95_99','years_100_plus')))


    df.columns = map(str.upper, df.columns)
    df.to_sql(RAW_CDC_85_PLUS_POPULATION_DATA_TABLE, sqlite_connection, if_exists='replace',index=False)
    print("Loaded 85 plus data in table:",RAW_CDC_85_PLUS_POPULATION_DATA_TABLE)

def update_85_plus_population_data():
    df = pd.read_sql_query(
        "SELECT YEAR,ETHNICITY,AGE_COHORT,SUM(POPULATION) as POPULATION from " + RAW_CDC_85_PLUS_POPULATION_DATA_TABLE + " WHERE ESTIMATE_MONTH=7 AND SERIES_VINTAGE=2019 group by YEAR,ETHNICITY,AGE_COHORT order by YEAR,ETHNICITY,AGE_COHORT",
        sqlite_connection)

    df = df.pivot_table(index=['ETHNICITY','YEAR'], columns='AGE_COHORT', values='POPULATION').reset_index()
    df.columns.name = None
    df = df.apply(pd.to_numeric, errors='ignore')

    for index, row in df.iterrows():
        sql="UPDATE "+POPULATION_DATA_TABLE+" SET years_85_89="+str(row['years_85_89'])+\
                                                 ",years_90_94="+str(row['years_90_94'])+\
                                                  ",years_95_99="+str(row['years_95_99'])+ \
                                                  ",years_100_plus=" + str(row['years_100_plus'])+" WHERE ETHNICITY='"+row['ETHNICITY']+"' AND YEAR="+str(row['YEAR'])

        sqlite_connection.execute(sql)
        print("Updated Data for Year and Ethnicity",row['YEAR'],row['ETHNICITY'])

    print("completed Updating 85 plus data")

def load_cdc_population_data(update_year=None):
    file=os.path.join(RAW_CDC_POPULATION_FILES, "cdc_raw_population_data.xlsx")
    population_xl = pd.ExcelFile(file)
    df_pop = population_xl.parse(sheet='2000-2019',nrows=3912)

    #drop total row
    df_pop=df_pop[(df_pop['Notes']!='Total') & (df_pop['Population']!='Not Applicable')]
    df_pop=df_pop.dropna(how='all')
    df_pop['ETHNICITY']=np.where( (df_pop['Gender']=='Male')&(df_pop['Race']=='White'),'EAM',
                                  np.where ( (df_pop['Gender']=='Female')&(df_pop['Race']=='White'),'EAF',
                                      np.where((df_pop['Gender'] == 'Male') & (df_pop['Race'] != 'White'), 'NEAM',
                                        'NEAF')))

    df_pop = df_pop.dropna(subset=['Gender'])
    df_pop=df_pop[['Five-Year Age Groups','ETHNICITY','Year','Population']]


    df_pop=df_pop.groupby(['Five-Year Age Groups','ETHNICITY','Year'])['Population'].sum()
    df_pop=df_pop.reset_index()

    df_pop = df_pop.pivot_table(index=['ETHNICITY','Year'], columns='Five-Year Age Groups', values='Population').reset_index()
    df_pop.columns.name = None
    COL_ORDER=['Year','ETHNICITY','< 1 year','1-4 years','5-9 years','10-14 years',	'15-19 years',	'20-24 years',	'25-29 years',	'30-34 years',	'35-39 years',	'40-44 years',	'45-49 years',	'50-54 years',	'55-59 years',		'60-64 years ',	'65-69 years',	'70-74 years',	'75-79 years',	'80-84 years']
    df_pop = df_pop[COL_ORDER]
    COL_ORDER=['YEAR','ETHNICITY','under_one_year', 'years_1_4','years_5_9', 'years_10_14', 'years_15_19', 'years_20_24', 'years_25_29', 'years_30_34', 'years_35_39', 'years_40_44', 'years_45_49', 'years_50_54', 'years_55_59', 'years_60_64', 'years_65_69', 'years_70_74', 'years_75_79', 'years_80_84']

    #change the columns names the way we want to store in the system
    df_pop.columns=COL_ORDER
    #as of now we don't have these data points
    df_pop[ 'years_85_89']=0
    df_pop['years_90_94'] = 0
    df_pop['years_95_99'] = 0
    df_pop['years_100_plus'] = 0
    COL_ORDER.extend(['years_85_89','years_90_94','years_95_99','years_100_plus'])


    SUM_COLS = [age for age in COL_ORDER if age not in ['YEAR','ETHNICITY', 'Not_stated']]
    df_pop['ALL_AGES'] = df_pop[SUM_COLS].sum(axis=1)


    #save CDC population
    df_pop.to_sql(RAW_CDC_POPULATION_DATA_TABLE, sqlite_connection, if_exists='replace', index=False)

    print("Saved CDC data in to ",RAW_CDC_POPULATION_DATA_TABLE)

    #update the Master population if year is provided

    if(update_year is not None):
        df_pop_update=df_pop[df_pop['YEAR']>=update_year]

        if(df_pop_update.shape[0] >0):
            df = pd.read_sql_query(
                "SELECT *  from " + POPULATION_DATA_TABLE +" WHERE YEAR>="+str(update_year),
                sqlite_connection)
            if(df.shape[0]>0): # if the data is already there then remove it

                sql = "DELETE FROM "+POPULATION_DATA_TABLE+" WHERE  YEAR>="+str(update_year)

                sqlite_connection.execute(sql)
                print("--Deleted exisitng data for Year >",update_year)

            df_pop_update.to_sql(POPULATION_DATA_TABLE, sqlite_connection, if_exists='append', index=False)

            print("---Updated Data in to  ", POPULATION_DATA_TABLE,' Starting Year:',str(update_year))



def load_tot_adjustment_factor(update_year=None):
    df_tot_eam=get_raw_mortality_data('All Causes')


    if(update_year is not None):
        df_tot_eam=df_tot_eam[df_tot_eam['YEAR']>=update_year]

    # df_tot_eam['years_1_4']=df_tot_eam['year_1']+df_tot_eam['years_2']+df_tot_eam['years_3']+df_tot_eam['years_4']
    df = df_tot_eam.melt(id_vars=['YEAR','ETHNICITY'], value_vars=df_tot_eam.columns.drop(['YEAR','CANCER_TYPE','ETHNICITY','TOTAL','under_5_years','Not_stated']).to_list(), var_name='AGE_COHORT', value_name='OBS')

    #get the population data
    df_pop = get_population_data()
    if(update_year is not None):
        df_pop=df_pop[df_pop['YEAR']>=update_year]

    df_pop = df_pop.apply(pd.to_numeric, errors='ignore')
    # we do not get these breakdown from CDC but need for Mortality Analysis
    df_pop['year_1'] = df_pop['years_1_4']/4
    df_pop['years_2'] = df_pop['years_1_4']/4
    df_pop['years_3'] = df_pop['years_1_4']/4
    df_pop['years_4'] = df_pop['years_1_4']/4

    #drop 1_4 years
    df_pop.drop('years_1_4', axis=1, inplace=True)

    df_pop = df_pop.melt(id_vars=['YEAR','ETHNICITY'], value_vars=df_pop.columns.drop(['YEAR','ALL_AGES', 'ETHNICITY']).to_list(), var_name='AGE_COHORT', value_name='POPULATION')

    df_tot = pd.merge(df, df_pop, on=['YEAR', 'AGE_COHORT','ETHNICITY'], how='left')

    df_tot = df_tot.apply(pd.to_numeric, errors='ignore')

    df_tot['TOT_ADJ_FACTOR']=1
    #if we have valid population numbers then calculate the ADJ Factor
    df_tot.loc[df_tot['POPULATION'] > 0, 'TOT_ADJ_FACTOR'] = 1 - (df_tot['OBS'] / df_tot['POPULATION'])

    if(df_tot['TOT_ADJ_FACTOR'].max() >1):
        sys.exit("1-TOT is greater than 1,so there is some issue with the code - Quitting Process !!!!!!!!!!!")

    df_tot_wide=df_tot[['YEAR','ETHNICITY','AGE_COHORT','TOT_ADJ_FACTOR']]
    df_tot_wide = df_tot_wide.pivot_table(index=['YEAR','ETHNICITY'], columns='AGE_COHORT', values='TOT_ADJ_FACTOR').reset_index()
    df_tot_wide.columns.name = None
    df_tot_wide['TOTAL']=np.NaN
    df_tot_wide['under_5_years'] = np.NaN




    return [df_tot,df_tot_wide]


def load_raw_adjustment_mortality(cancer_type,df_tot_adjustment,update_year=None):
    df_tot_eam = get_raw_mortality_data(cancer_type)

    if (update_year is not None):
        df_tot_eam = df_tot_eam[df_tot_eam['YEAR'] >= update_year]

    # df_tot_eam['years_1_4']=df_tot_eam['year_1']+df_tot_eam['years_2']+df_tot_eam['years_3']+df_tot_eam['years_4']
    df = df_tot_eam.melt(id_vars=['YEAR', 'ETHNICITY'], value_vars=df_tot_eam.columns.drop(['YEAR', 'CANCER_TYPE', 'ETHNICITY', 'TOTAL', 'under_5_years', 'Not_stated']).to_list(), var_name='AGE_COHORT', value_name='MOR')

    df_tot_adjustment=df_tot_adjustment[['YEAR', 'AGE_COHORT', 'ETHNICITY','TOT_ADJ_FACTOR','POPULATION']]
    df_tot = pd.merge(df, df_tot_adjustment, on=['YEAR', 'AGE_COHORT', 'ETHNICITY'], how='left')

    df_tot = df_tot.apply(pd.to_numeric, errors='ignore')

    #define OBS
    df_tot['OBS'] = df_tot['MOR']/df_tot['POPULATION']
    df_tot.loc[~np.isfinite(df_tot['OBS']), 'OBS'] = np.nan

    #if 1-TOT is <0 then we have issue with population breakdown. Just make it to 1 otherwise we will get negative number
    df_tot.loc[df_tot['TOT_ADJ_FACTOR']<0,'TOT_ADJ_FACTOR']=1

    if(df_tot['TOT_ADJ_FACTOR'].max() >1):
        sys.exit("1-TOT is greater than 1,so there is some issue with the code - Quitting Process !!!!!!!!!!!")

    #caclulate OBS*
    df_tot['RAW_ADJ'] = df_tot['MOR']/(df_tot['TOT_ADJ_FACTOR']+df_tot['OBS'])
    #df_tot['OBS_STAR'] = (df_tot['OBS']) / (df_tot['TOT_ADJ_FACTOR'] + df_tot['OBS'])


    df_tot_wide = df_tot[['YEAR', 'ETHNICITY', 'AGE_COHORT', 'RAW_ADJ']]
    df_tot_wide = df_tot_wide.pivot_table(index=['YEAR', 'ETHNICITY'], columns='AGE_COHORT', values='RAW_ADJ').reset_index()
    df_tot_wide.columns.name = None
    df_tot_wide['under_5_years']=df_tot_wide[['under_one_year','year_1','years_2','years_3','years_4']].sum(axis=1)
    df_tot_wide['Not_stated']=0
    df_tot_wide['TOTAL'] = df_tot_wide[AGE_OF_DEATH_SUM_COLS].sum(axis=1)

    df_OBS_wide = df_tot[['YEAR', 'ETHNICITY', 'AGE_COHORT', 'OBS']]
    df_OBS_wide = df_OBS_wide.pivot_table(index=['YEAR', 'ETHNICITY'], columns='AGE_COHORT', values='OBS').reset_index()
    df_OBS_wide.columns.name = None
    df_OBS_wide['under_5_years'] = df_OBS_wide[['under_one_year', 'year_1', 'years_2', 'years_3', 'years_4']].sum(axis=1)
    df_OBS_wide['Not_stated'] = 0
    df_OBS_wide['TOTAL'] = df_OBS_wide[AGE_OF_DEATH_SUM_COLS].sum(axis=1)


    df_tot['CANCER_TYPE']=cancer_type

    return [df_tot,df_tot_wide,df_OBS_wide]


def load_birth_year_cohort_mortality(cancer_type,df,by_birth_year_cohort=1,bForCharts=False):

    if(bForCharts):
        df['AGE_AT_DEATH'] = df['AGE_COHORT'].map(age_of_death_dict_for_charts)

    else:
        df['AGE_AT_DEATH'] = df['AGE_COHORT'].map(age_of_death_dict)

    df['YEAR_OF_BIRTH'] = (df['YEAR'] - df['AGE_AT_DEATH'])
    # df['YEAR_OF_BIRTH'] = np.ceil(df['YEAR_OF_BIRTH']).astype(int)
    # df['YEAR_OF_BIRTH'] = np.floor(df['YEAR_OF_BIRTH']).astype(int)

    df['YEAR_OF_BIRTH'] =np.where( ((df['AGE_AT_DEATH'] == 0.5)), np.floor(df['YEAR_OF_BIRTH']).astype(int),np.ceil(df['YEAR_OF_BIRTH']).astype(int))

    df.drop('AGE_COHORT', axis=1, inplace=True)

    df = df.apply(pd.to_numeric, errors='ignore')
    df=df[['ETHNICITY','AGE_AT_DEATH','YEAR_OF_BIRTH','RAW_ADJ','POPULATION']]

    #df['OBS_BY_BIRTH_YEAR']=(df['RAW_ADJUSTED_MORTALITY']/df['POPULATION'])*(10**5)

    df = df.sort_values(by=['AGE_AT_DEATH', 'YEAR_OF_BIRTH'])

    # df['DECADE_OF_BIRTH'] = (df['YEAR_OF_BIRTH'] // by_birth_year_cohort) * by_birth_year_cohort
    df['YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort)] = (df['YEAR_OF_BIRTH'] // by_birth_year_cohort) * by_birth_year_cohort



    #df.drop(['POPULATION', 'RAW_ADJUSTED_MORTALITY','YEAR_OF_BIRTH'], axis=1, inplace=True)
    df = df.groupby(['ETHNICITY','AGE_AT_DEATH', 'YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort)])['POPULATION','RAW_ADJ'].sum()
    df=df.reset_index()

    # df['OBS_BY_BIRTH_YEAR']=(df['RAW_ADJ']/df['POPULATION'])*(10**5)
    #removed 10**5  asper advice from Prof Thilly
    df['OBS_BY_BIRTH_YEAR']=(df['RAW_ADJ']/df['POPULATION'])

    df.loc[~np.isfinite(df['OBS_BY_BIRTH_YEAR']), 'OBS_BY_BIRTH_YEAR'] = np.nan

    df = df.astype({'YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort): int})

    if(bForCharts):
        df_wide = df.pivot_table(index=['YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort), 'ETHNICITY'], columns='AGE_AT_DEATH', values='OBS_BY_BIRTH_YEAR').reset_index()
        df_wide.columns.name = None
    else:
        df_wide = df.pivot_table(index=['AGE_AT_DEATH','ETHNICITY'], columns='YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort), values='OBS_BY_BIRTH_YEAR').reset_index()
        df_wide.columns.name = None

    #save this tall dataframe in the sqlite table
    df['CANCER_TYPE']=cancer_type
    df['YEAR_OF_BIRTH_COHORT'] = df['YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort)]
    df.drop('YEAR_OF_BIRTH_COHORT_'+str(by_birth_year_cohort), axis=1, inplace=True)
    df['YEAR_OF_BIRTH_COHORT_TYPE'] = by_birth_year_cohort

    insert_birth_year_cohort_mortality(df,cancer_type,by_birth_year_cohort)




    return df_wide

def insert_birth_year_cohort_mortality(df_data,cancer_type,birth_year_cohort):

    # check if table exists
    df_table = pd.read_sql_query(
        "SELECT  name  from sqlite_master where type='table' AND name='" + MORTALITY_BY_BIRTH_YEAR_TABLE + "'",
        sqlite_connection)

    if(df_table.shape[0] >0):
        if ("'" in cancer_type):
            cancer_type = cancer_type.replace("'", "''")

        query="SELECT *  from " + MORTALITY_BY_BIRTH_YEAR_TABLE+" WHERE CANCER_TYPE='"+cancer_type+"' AND YEAR_OF_BIRTH_COHORT_TYPE="+str(birth_year_cohort)
        print(query)
        df = pd.read_sql_query(
            "SELECT *  from " + MORTALITY_BY_BIRTH_YEAR_TABLE+" WHERE CANCER_TYPE='"+cancer_type+"' AND YEAR_OF_BIRTH_COHORT_TYPE="+str(birth_year_cohort),
            sqlite_connection)
        if(df.shape[0]>0): # if the dat is already there then remove it
            print('Deleting existing Data for cancer type',cancer_type,' and birth year cohort',birth_year_cohort,' from ',MORTALITY_BY_BIRTH_YEAR_TABLE)

            sql = "DELETE FROM "+MORTALITY_BY_BIRTH_YEAR_TABLE+" WHERE CANCER_TYPE='"+cancer_type+"' AND YEAR_OF_BIRTH_COHORT_TYPE="+str(birth_year_cohort)

            sqlite_connection.execute(sql)

    #append the new data
    df_data.to_sql(MORTALITY_BY_BIRTH_YEAR_TABLE, sqlite_connection, if_exists='append', index=False)





def generate_txt_files(df,cancer_type,birth_year_cohort,ethnicity,cancer_type_txt_file):
    # if ("'" in cancer_type):
    #     cancer_type = cancer_type.replace("'", "''")
    #
    # df = pd.read_sql_query(
    #     "SELECT *  from " + MORTALITY_BY_BIRTH_YEAR_TABLE+" WHERE OBS_BY_BIRTH_YEAR>0 AND CANCER_TYPE='"+cancer_type+"' AND YEAR_OF_BIRTH_COHORT_TYPE="+str(birth_year_cohort)+" AND ETHNICITY='"+ethnicity+"' order by YEAR_OF_BIRTH_COHORT,AGE_AT_DEATH ASC",
    #         sqlite_connection)
    #
    # if(df.shape[0]==0):
    #     print(" NO DATA TO GENERATE TXT FILES FOR CANCER:",cancer_type,birth_year_cohort,ethnicity)
    #     return
    # df_csv=df.pivot_table(index=['YEAR_OF_BIRTH_COHORT','ETHNICITY'], columns='AGE_AT_DEATH', values='OBS_BY_BIRTH_YEAR').reset_index()
    # df_csv.columns.name = None
    df_csv=df[df['ETHNICITY']==ethnicity]
    df_csv.drop('ETHNICITY', axis=1, inplace=True)

    if (birth_year_cohort==10):
        txt_file_name = cancer_type_txt_file + "_" + ethnicity + ".txt"
        data_file_name = os.path.join(DECADES_MORTALITY_DATA_TXT_FILES, txt_file_name)
        try:
            os.remove(data_file_name)
        except OSError:
            pass
        df_csv.to_csv(data_file_name,header=None, index=False, sep='\t')
        txt_file_name2 = cancer_type_txt_file + "_" + ethnicity + "_y_" + str(birth_year_cohort) + ".txt"
        data_file_name = os.path.join(DECADES_MORTALITY_DATA_TXT_FILES, txt_file_name2)
        try:
            os.remove(data_file_name)
        except OSError:
            pass
        df_csv.to_csv(data_file_name, header=None, index=False, sep='\t')
    if (birth_year_cohort==5):
        txt_file_name = cancer_type_txt_file + "_" + ethnicity + "_y_" + str(birth_year_cohort) + ".txt"
        data_file_name = os.path.join(DECADES_MORTALITY_DATA_TXT_FILES, txt_file_name)
        try:
            os.remove(data_file_name)
        except OSError:
            pass
        df_csv.to_csv(data_file_name, header=None, index=False, sep='\t')
    print("Comepleted generating TXT files for cancer:",cancer_type," file name",data_file_name)



#=======================================================================================================

if __name__ == "__main__":

    #Step 1

    # #load raw mortality data for a given year - These files are provided by Karl Rexer
    # load_ucr_files(2019)
    # YEARS=[2011,2012,2013,2014,2015,2016]
    # load_ICD10_files()

    #Step 2 - This is one time task to load the UCR mapping
    # # load_ucr358_codes("2011_UCR358_CODES",UCR358_CODES_TABLE)

    # Step4:
    # Load the population and TOT Adjustment factor
    # load_master_population_data()

    #load the population data from CDC -
    # load_cdc_population_data(2015)
    #===============================================================
    #load 85 plus population data and update the master population data table
    # load_85_plus_population_data()
    # update_85_plus_population_data()
    #$===============================================


    #Step 3:
    # PATH=(base_path/ "ucr_output_data_files/").resolve()
    # # PATH = (base_path / "test_folder/").resolve()
    # load_existing_mortality_raw_data(PATH)


    #======Testing only==========================
    # df_tot,df_tot_wide=load_tot_adjustment_factor(1900)
    # #
    # cancer_type='Infectious & Parasitic Diseases'
    # df_adj,df_adj_wide=load_raw_adjustment_mortality(cancer_type,df_tot)
    # df_adj_10=df_adj.copy()
    # df_10_year=load_birth_year_cohort_mortality(cancer_type,df_adj_10,10,True)
    # df_5_year=load_birth_year_cohort_mortality(cancer_type,df_adj.copy(),5,True)
    # df_1_year = load_birth_year_cohort_mortality(cancer_type,df_adj.copy(), 1,True)
    # generate_txt_files(cancer_type,10,'EAF')
    # generate_txt_files(cancer_type,5,'EAF')
    # generate_txt_files(cancer_type,10,'EAM')
    # generate_txt_files(cancer_type,5,'EAM')
    # generate_txt_files(cancer_type,10,'NEAF')
    # generate_txt_files(cancer_type,5,'NEAF')
    # generate_txt_files(cancer_type,10,'NEAM')
    # generate_txt_files(cancer_type,5,'NEAM')



    # print(df_adj.head())
    # ======Testing only=========================





    #load the original data fo all cancer types from 2007 to 2010. We can compare this data with the newly loaded data to check our UCR358 mapping
    # load_existing_data_for_testing()



    # # df=get_ucr358_codes()
    # # print(df.head())
    # # df=get_mit_ucr358_mapping()
    #


    ##no issues running everytime..table will br dropped and recreated everytime
    load_ucr358_codes("MIT_UCR358", MIT_UCR358_CODE_MAPPING_TABLE)
    # #generate MIT Data files
    generate_mit_data_files()

    #load the population data from All causes xl file  and load 1-Tot data from same file
    # load_existing_population_data()



    #this creates a unified data rectangle combining all Raw data and Adjusted raw data
    # create_mit_data_rectangle(OUTPUT_DATA_FILE_PATH)







